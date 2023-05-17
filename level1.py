import unittest
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import NoSuchElementException

from openpyxl import load_workbook


class Utils:
    def login(driver):
        element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.LINK_TEXT, "Log in"))
        )

        element.click()
        user = driver.find_element("name", "username")
        user.send_keys("teacher")
        password = driver.find_element("name", "password")
        password.send_keys("moodle")
        password.send_keys(Keys.RETURN)

    def read_data_from_excel(filename, sheet):
        datalist = []
        wb = load_workbook(filename=filename)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(4, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        print(datalist)
        return datalist


class Grading(unittest.TestCase):
    def __init__(self, test_method_name, test_data):
        super().__init__(test_method_name)
        self.test_data = test_data

    def setUp(self):
        PATH = "D:\Download\chromedriver_win32\chromedriver.exe"

        self.driver = webdriver.Chrome(PATH)
        self.driver.maximize_window()
        self.driver.get(
            "https://school.moodledemo.net/mod/assign/view.php?id=980&action=grader&userid=10")
        Utils.login(self.driver)

    def test_grading(self):
        driver = self.driver

        driver.get(
            "https://school.moodledemo.net/mod/assign/view.php?id=980&action=grader&userid=10")

        test_id, grade_ex, action, output, expected_url = self.test_data

        driver = self.driver
        time.sleep(10)
        driver.find_element(
            By.XPATH, "//div[3]/div/div[2]/div/div[2]/input").click()
        grade = driver.find_element("name", "grade")
        grade.clear()
        if grade_ex is not None:
            grade.send_keys(grade_ex)

        if action == 'Save Changes':
            save = driver.find_element("name", "savechanges")
            save.click()
        if action == 'Save and show next':
            save = driver.find_element("name", "saveandshownext")
            save.click()
        time.sleep(15)
        driver.find_element(
            By.XPATH, "//div[3]/div/div[2]/div/div[2]/input").click()
        time.sleep(5)

        # Check if grade setting was successful
        if output is None:
            self.assertEqual(driver.current_url, expected_url)
        else:
            error_list = output

            # Check if error message is displayed
            try:
                error = driver.find_element(
                    by=By.XPATH, value="//fieldset/div[2]/div/div[2]/div"
                )
            except NoSuchElementException:
                self.fail(f"Test {test_id} failed")
            if error.text not in error_list:
                self.fail(f"Test {test_id} failed")
            self.assertEqual(driver.current_url, expected_url)

    def tearDown(self):
        self.driver.close()


if __name__ == "__main__":
    # Read test data from Excel file
    testcases = Utils.read_data_from_excel("./testcase.xlsx", "level1")

    # Create a test suite and add the GradeSettingTest class
    suite = unittest.TestSuite()

    # Loop through test cases and add a test case for each one
    for test_data in testcases:
        test_case = Grading(
            "test_grading", test_data=test_data)
        suite.addTest(test_case)

    # Run the test suite
    runner = unittest.TextTestRunner(verbosity=2)
    runner.run(suite)
